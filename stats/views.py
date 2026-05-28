from django.shortcuts import render, redirect
from django.http import JsonResponse, HttpResponse
from django.contrib import messages
from django.utils import timezone
from datetime import timedelta, date
from django.db.models import Count, Sum, Q
from borrow.models import BorrowRecord, FineRecord, OperationLog, SystemConfig
from books.models import Book, Category
from django.contrib.auth import get_user_model
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO
from django.contrib.auth.decorators import login_required
import json

User = get_user_model()


def _is_ajax(request):
    return request.headers.get("X-Requested-With") == "XMLHttpRequest" or request.META.get("HTTP_X_REQUESTED_WITH") == "XMLHttpRequest"


def _json_response(code, msg, data=None):
    payload = {"code": code, "msg": msg}
    if data is not None:
        payload["data"] = data
    return JsonResponse(payload)


def _admin_required(request):
    if request.user.role != "admin":
        if _is_ajax(request):
            return _json_response(403, "您没有管理员权限")
        messages.error(request, "您没有管理员权限")
        return redirect("/")
    return None


@login_required
def stats_dashboard_view(request):
    admin_check = _admin_required(request)
    if admin_check:
        return admin_check

    try:
        now = timezone.now()
        today_start = now.replace(hour=0, minute=0, second=0, microsecond=0)
        week_start = today_start - timedelta(days=7)
        month_start = today_start - timedelta(days=30)

        today_borrow_count = BorrowRecord.objects.filter(borrow_date__gte=today_start).count()
        today_return_count = BorrowRecord.objects.filter(return_date__gte=today_start).count()
        today_overdue_count = BorrowRecord.objects.filter(status="overdue").count()

        week_borrow_count = BorrowRecord.objects.filter(borrow_date__gte=week_start).count()
        week_return_count = BorrowRecord.objects.filter(return_date__gte=week_start).count()
        week_overdue_count = BorrowRecord.objects.filter(status="overdue").count()

        month_borrow_count = BorrowRecord.objects.filter(borrow_date__gte=month_start).count()
        month_return_count = BorrowRecord.objects.filter(return_date__gte=month_start).count()
        month_overdue_count = BorrowRecord.objects.filter(status="overdue").count()

        total_books = Book.objects.filter(is_deleted=False).count()
        total_users = User.objects.count()
        total_borrows = BorrowRecord.objects.count()

        category_stats = Category.objects.annotate(
            book_count=Count("book", filter=Q(book__is_deleted=False)),
            total_stock=Sum("book__total_stock"),
            current_stock=Sum("book__current_stock"),
        ).order_by("name")

        monthly_trends = []
        for i in range(5, -1, -1):
            month_date = today_start.replace(day=1) - timedelta(days=i * 30)
            month_date = month_date.replace(day=1)
            if i == 0:
                next_month = today_start.replace(day=28) + timedelta(days=4)
                month_end = next_month - timedelta(days=next_month.day)
            else:
                next_month = (month_date.replace(day=28) + timedelta(days=4))
                month_end = next_month - timedelta(days=next_month.day)
            month_label = month_date.strftime("%Y-%m")
            b_cnt = BorrowRecord.objects.filter(
                borrow_date__gte=month_date, borrow_date__lte=month_end
            ).count()
            r_cnt = BorrowRecord.objects.filter(
                return_date__gte=month_date, return_date__lte=month_end
            ).count()
            monthly_trends.append({
                "label": month_label,
                "borrow": b_cnt,
                "return": r_cnt,
            })

        daily_trends = []
        for i in range(13, -1, -1):
            day_start = today_start - timedelta(days=i)
            day_end = day_start + timedelta(days=1)
            day_label = day_start.strftime("%m-%d")
            d_cnt = BorrowRecord.objects.filter(
                borrow_date__gte=day_start, borrow_date__lt=day_end
            ).count()
            daily_trends.append({"label": day_label, "count": d_cnt})

        borrow_status_stats = [
            {"name": "借阅中", "value": BorrowRecord.objects.filter(status="borrowing").count()},
            {"name": "已归还", "value": BorrowRecord.objects.filter(status="returned").count()},
            {"name": "已逾期", "value": BorrowRecord.objects.filter(status="overdue").count()},
        ]

        category_chart_data = [
            {"name": c.name, "value": c.book_count}
            for c in category_stats if c.book_count and c.book_count > 0
        ]

        top_users = User.objects.annotate(
            br_count=Count("borrowrecord")
        ).order_by("-br_count")[:8]

        user_borrow_stats = [
            {"name": u.username, "value": u.br_count}
            for u in top_users if u.br_count > 0
        ]

        context = {
            "today_borrow_count": today_borrow_count,
            "today_return_count": today_return_count,
            "today_overdue_count": today_overdue_count,
            "week_borrow_count": week_borrow_count,
            "week_return_count": week_return_count,
            "week_overdue_count": week_overdue_count,
            "month_borrow_count": month_borrow_count,
            "month_return_count": month_return_count,
            "month_overdue_count": month_overdue_count,
            "total_books": total_books,
            "total_users": total_users,
            "total_borrows": total_borrows,
            "category_stats": category_stats,
            "monthly_trends_json": json.dumps(monthly_trends, ensure_ascii=False),
            "daily_trends_json": json.dumps(daily_trends, ensure_ascii=False),
            "borrow_status_json": json.dumps(borrow_status_stats, ensure_ascii=False),
            "category_chart_json": json.dumps(category_chart_data, ensure_ascii=False),
            "user_borrow_json": json.dumps(user_borrow_stats, ensure_ascii=False),
        }
        return render(request, "stats/dashboard.html", context)

    except Exception as e:
        messages.error(request, f"加载统计面板失败: {str(e)}")
        return render(request, "stats/dashboard.html", {
            "today_borrow_count": 0,
            "today_return_count": 0,
            "today_overdue_count": 0,
            "week_borrow_count": 0,
            "week_return_count": 0,
            "week_overdue_count": 0,
            "month_borrow_count": 0,
            "month_return_count": 0,
            "month_overdue_count": 0,
            "total_books": 0,
            "total_users": 0,
            "total_borrows": 0,
            "category_stats": [],
            "monthly_trends_json": "[]",
            "daily_trends_json": "[]",
            "borrow_status_json": "[]",
            "category_chart_json": "[]",
            "user_borrow_json": "[]",
        })


@login_required
def stats_hot_books_view(request):
    admin_check = _admin_required(request)
    if admin_check:
        return admin_check

    try:
        hot_books = Book.objects.filter(is_deleted=False).annotate(
            borrow_count_anno=Count("borrowrecord")
        ).order_by("-borrow_count_anno")[:10]

        context = {"hot_books": hot_books}
        return render(request, "stats/hot_books.html", context)

    except Exception as e:
        messages.error(request, f"加载热门图书失败: {str(e)}")
        return render(request, "stats/hot_books.html", {"hot_books": []})


@login_required
def stats_inventory_view(request):
    admin_check = _admin_required(request)
    if admin_check:
        return admin_check

    try:
        category_stats = Category.objects.annotate(
            book_count=Count("book", filter=Q(book__is_deleted=False)),
            total_stock=Sum("book__total_stock"),
            current_stock=Sum("book__current_stock"),
        ).order_by("name")

        out_of_stock_books = Book.objects.filter(is_deleted=False, current_stock=0).select_related("category").order_by("title")

        context = {
            "category_stats": category_stats,
            "out_of_stock_books": out_of_stock_books,
        }
        return render(request, "stats/inventory.html", context)

    except Exception as e:
        messages.error(request, f"加载库存统计失败: {str(e)}")
        return render(request, "stats/inventory.html", {
            "category_stats": [],
            "out_of_stock_books": [],
        })


def _apply_excel_style(ws, headers, col_widths=None):
    header_font = Font(name="微软雅黑", bold=True, size=11)
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border

    if col_widths:
        for col_idx, width in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = width


@login_required
def stats_export_borrows_view(request):
    admin_check = _admin_required(request)
    if admin_check:
        return admin_check

    try:
        queryset = BorrowRecord.objects.select_related("user", "book").all()

        status_filter = request.GET.get("status", "").strip()
        if status_filter:
            queryset = queryset.filter(status=status_filter)

        start_date = request.GET.get("start_date", "").strip()
        end_date = request.GET.get("end_date", "").strip()
        if start_date:
            queryset = queryset.filter(borrow_date__gte=start_date)
        if end_date:
            queryset = queryset.filter(borrow_date__lte=end_date + " 23:59:59")

        queryset = queryset.order_by("-borrow_date")

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "借阅记录"

        headers = ["编号", "用户名", "书名", "ISBN", "借阅日期", "应还日期", "归还日期", "状态", "罚款"]
        col_widths = [8, 16, 30, 18, 20, 20, 20, 10, 10]
        _apply_excel_style(ws, headers, col_widths)

        data_font = Font(name="微软雅黑", size=10)
        data_alignment = Alignment(vertical="center")
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        status_map = {"borrowing": "借阅中", "returned": "已归还", "overdue": "已逾期"}

        for row_idx, record in enumerate(queryset, 2):
            ws.cell(row=row_idx, column=1, value=record.id).font = data_font
            ws.cell(row=row_idx, column=1).alignment = data_alignment
            ws.cell(row=row_idx, column=1).border = thin_border

            ws.cell(row=row_idx, column=2, value=record.user.username).font = data_font
            ws.cell(row=row_idx, column=2).alignment = data_alignment
            ws.cell(row=row_idx, column=2).border = thin_border

            ws.cell(row=row_idx, column=3, value=record.book.title).font = data_font
            ws.cell(row=row_idx, column=3).alignment = data_alignment
            ws.cell(row=row_idx, column=3).border = thin_border

            ws.cell(row=row_idx, column=4, value=record.book.isbn).font = data_font
            ws.cell(row=row_idx, column=4).alignment = data_alignment
            ws.cell(row=row_idx, column=4).border = thin_border

            ws.cell(row=row_idx, column=5, value=record.borrow_date.strftime("%Y-%m-%d %H:%M") if record.borrow_date else "").font = data_font
            ws.cell(row=row_idx, column=5).alignment = data_alignment
            ws.cell(row=row_idx, column=5).border = thin_border

            ws.cell(row=row_idx, column=6, value=record.due_date.strftime("%Y-%m-%d %H:%M") if record.due_date else "").font = data_font
            ws.cell(row=row_idx, column=6).alignment = data_alignment
            ws.cell(row=row_idx, column=6).border = thin_border

            ws.cell(row=row_idx, column=7, value=record.return_date.strftime("%Y-%m-%d %H:%M") if record.return_date else "").font = data_font
            ws.cell(row=row_idx, column=7).alignment = data_alignment
            ws.cell(row=row_idx, column=7).border = thin_border

            ws.cell(row=row_idx, column=8, value=status_map.get(record.status, record.status)).font = data_font
            ws.cell(row=row_idx, column=8).alignment = data_alignment
            ws.cell(row=row_idx, column=8).border = thin_border

            ws.cell(row=row_idx, column=9, value=float(record.fine_amount)).font = data_font
            ws.cell(row=row_idx, column=9).alignment = data_alignment
            ws.cell(row=row_idx, column=9).border = thin_border

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(
            output.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = 'attachment; filename="borrow_records.xlsx"'
        return response

    except Exception as e:
        messages.error(request, f"导出借阅记录失败: {str(e)}")
        return redirect("stats_dashboard")


@login_required
def stats_export_books_view(request):
    admin_check = _admin_required(request)
    if admin_check:
        return admin_check

    try:
        books = Book.objects.filter(is_deleted=False).select_related("category").order_by("-created_at")

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "图书库存"

        headers = ["编号", "书名", "作者", "ISBN", "分类", "总库存", "当前库存", "借阅次数"]
        col_widths = [8, 30, 16, 18, 14, 10, 10, 10]
        _apply_excel_style(ws, headers, col_widths)

        data_font = Font(name="微软雅黑", size=10)
        data_alignment = Alignment(vertical="center")
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        for row_idx, book in enumerate(books, 2):
            row_data = [
                book.id,
                book.title,
                book.author,
                book.isbn,
                book.category.name if book.category else "",
                book.total_stock,
                book.current_stock,
                book.borrow_count,
            ]
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = data_font
                cell.alignment = data_alignment
                cell.border = thin_border

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(
            output.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = 'attachment; filename="books_inventory.xlsx"'
        return response

    except Exception as e:
        messages.error(request, f"导出图书库存失败: {str(e)}")
        return redirect("stats_dashboard")


@login_required
def stats_export_users_view(request):
    admin_check = _admin_required(request)
    if admin_check:
        return admin_check

    try:
        users = User.objects.all().order_by("-date_joined")

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "用户数据"

        headers = ["编号", "用户名", "手机号", "角色", "借阅数量", "累计罚款", "注册时间", "状态"]
        col_widths = [8, 16, 14, 10, 10, 10, 20, 10]
        _apply_excel_style(ws, headers, col_widths)

        data_font = Font(name="微软雅黑", size=10)
        data_alignment = Alignment(vertical="center")
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        role_map = {"user": "普通用户", "admin": "管理员"}

        for row_idx, user in enumerate(users, 2):
            row_data = [
                user.id,
                user.username,
                user.phone or "",
                role_map.get(user.role, user.role),
                user.borrow_count,
                float(user.total_fine),
                user.date_joined.strftime("%Y-%m-%d %H:%M") if user.date_joined else "",
                "已冻结" if user.is_frozen else "正常",
            ]
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = data_font
                cell.alignment = data_alignment
                cell.border = thin_border

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(
            output.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = 'attachment; filename="users_data.xlsx"'
        return response

    except Exception as e:
        messages.error(request, f"导出用户数据失败: {str(e)}")
        return redirect("stats_dashboard")